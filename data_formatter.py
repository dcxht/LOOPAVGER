"""
Respiratory Data Processor
Data Formatter Interface - Converts raw respiratory data files to a standard format
Includes batch processing capability
"""

import tkinter as tk
from tkinter import Label, Entry, Button, StringVar, filedialog, messagebox, Frame
from tkinter import Listbox, Scrollbar, END
import threading
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook

class DataFormatterInterface:
    """Interface for the Data Formatter functionality"""
    
    def __init__(self, parent):
        """
        Initialize the Data Formatter interface
        
        Parameters:
        parent - Parent tkinter window
        """
        self.parent = parent
        self.window = tk.Toplevel(parent)
        self.window.title("Data Formatter")
        self.window.geometry("600x500")
        self.window.resizable(True, True)
        
        # Data variables
        self.selected_files = []
        self.output_dir = StringVar()
        self.status_var = StringVar(value="Ready")
        
        # Create UI elements
        self.create_widgets()
        
        # Make window modal
        self.window.transient(parent)
        self.window.grab_set()
        
    def create_widgets(self):
        """Create and arrange all UI widgets"""
        main_frame = Frame(self.window, padx=20, pady=20)
        main_frame.pack(fill='both', expand=True)
        
        # Title
        title_label = Label(main_frame, text="Data Formatter", font=("Arial", 14, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        subtitle_label = Label(main_frame, 
                              text="Convert raw respiratory data files to the standard format", 
                              font=("Arial", 9, "italic"))
        subtitle_label.grid(row=0, column=0, columnspan=3, pady=(20, 0))
        
        # Files selection frame
        file_frame = Frame(main_frame)
        file_frame.grid(row=1, column=0, columnspan=3, pady=10, sticky="ew")
        
        file_label = Label(file_frame, text="Input Files:")
        file_label.grid(row=0, column=0, sticky="w")
        
        # File list with scrollbar
        list_frame = Frame(file_frame)
        list_frame.grid(row=1, column=0, columnspan=2, sticky="ew")
        
        scrollbar = Scrollbar(list_frame)
        scrollbar.pack(side="right", fill="y")
        
        self.file_listbox = Listbox(list_frame, height=8, width=60, selectmode=tk.MULTIPLE, yscrollcommand=scrollbar.set)
        self.file_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.file_listbox.yview)
        
        # File buttons frame
        btn_frame = Frame(file_frame)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=5, sticky="w")
        
        add_file_btn = Button(btn_frame, text="Add Files...", command=self.add_files)
        add_file_btn.pack(side="left", padx=5)
        
        clear_files_btn = Button(btn_frame, text="Clear Selection", command=self.clear_files)
        clear_files_btn.pack(side="left", padx=5)
        
        # Output directory selection
        output_label = Label(main_frame, text="Output Directory:", anchor="w")
        output_label.grid(row=2, column=0, sticky="w", pady=5)
        
        output_path_label = Label(main_frame, textvariable=self.output_dir, width=40, 
                                anchor="w", bg="#f0f0f0", relief="sunken", padx=5)
        output_path_label.grid(row=2, column=1, pady=5, sticky="w")
        
        output_btn = Button(main_frame, text="Browse...", command=self.browse_output_dir)
        output_btn.grid(row=2, column=2, padx=5)
        
        # Instructions
        instruction_frame = Frame(main_frame, bd=1, relief="solid", padx=10, pady=10)
        instruction_frame.grid(row=3, column=0, columnspan=3, sticky="ew", pady=15)
        
        instructions = Label(instruction_frame, text=(
            "Data Formatter converts raw respiratory data files to a standard format by:\n"
            "1. Identifying and extracting flow (ltr/s) and volume (ltr) data columns\n"
            "2. Creating standardized time intervals (0.01s)\n"
            "3. Generating properly formatted Excel files with Time, Vol, and Flow columns\n\n"
            "Batch processing: Select multiple files to convert them all at once.\n"
            "Output files will be saved in the selected output directory with '_formatted' suffix."
        ), justify="left")
        instructions.pack(anchor="w")
        
        # Button frame
        button_frame = Frame(main_frame)
        button_frame.grid(row=5, column=0, columnspan=3, pady=10)
        
        # Process button
        process_button = Button(button_frame, text="Process Files", command=self.start_processing, 
                               width=15, bg="#4CAF50", fg="white", font=("Arial", 10, "bold"))
        process_button.pack(side="left", padx=10)
        
        # Generate Plots button - disabled initially
        self.plot_button = Button(button_frame, text="Generate Plots", command=self.generate_plots,
                               width=15, bg="#4C75AF", fg="white", font=("Arial", 10), state="disabled")
        self.plot_button.pack(side="left", padx=10)
        
        # Close button
        close_button = Button(button_frame, text="Close", command=self.window.destroy, width=10)
        close_button.pack(side="left", padx=10)
        
        # Status bar
        status_bar = Label(main_frame, textvariable=self.status_var, bd=1, relief="sunken", anchor="w")
        status_bar.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(10, 0))
    
    def add_files(self):
        """Open file dialog to select multiple Excel files"""
        files = filedialog.askopenfilenames(
            title="Select Input Excel Files",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if files:
            for file_path in files:
                if file_path not in self.selected_files:
                    self.selected_files.append(file_path)
                    # Show just the file name in the listbox, not the full path
                    self.file_listbox.insert(END, os.path.basename(file_path))
    
    def clear_files(self):
        """Clear all selected files"""
        self.file_listbox.delete(0, END)
        self.selected_files = []
    
    def browse_output_dir(self):
        """Open file dialog to select output directory"""
        dir_path = filedialog.askdirectory(
            title="Select Output Directory"
        )
        
        if dir_path:
            self.output_dir.set(dir_path)
    
    def process_files(self):
        """Process all selected files"""
        # Validate input
        if not self.selected_files:
            messagebox.showerror("Error", "Please select at least one input file.")
            return
        
        if not self.output_dir.get():
            messagebox.showerror("Error", "Please specify an output directory.")
            return
        
        # Start processing in a separate thread
        self.status_var.set("Converting files...")
        thread = threading.Thread(
            target=self.run_conversion_thread
        )
        thread.daemon = True
        thread.start()
    
    def run_conversion_thread(self):
        """Run the conversion in a separate thread"""
        try:
            # Convert each file
            successful_files = []
            failed_files = []
            
            for input_file in self.selected_files:
                try:
                    # Generate output filename
                    base_name = os.path.splitext(os.path.basename(input_file))[0]
                    output_file = os.path.join(self.output_dir.get(), f"{base_name}_formatted.xlsx")
                    
                    # Process the file
                    self.convert_unedited_file(input_file, output_file)
                    successful_files.append(os.path.basename(input_file))
                except Exception as e:
                    failed_files.append(f"{os.path.basename(input_file)} (Error: {str(e)})")
            
            # Update UI
            self.window.after(0, lambda: self.batch_conversion_complete(successful_files, failed_files))
            
        except Exception as e:
            # Handle any exceptions
            self.window.after(0, lambda: messagebox.showerror(
                "Error", f"An error occurred during conversion: {str(e)}"
            ))
            self.window.after(0, lambda: self.status_var.set("Error during conversion."))
    
    def batch_conversion_complete(self, successful_files, failed_files):
        """Update the UI after batch conversion is complete"""
        if failed_files:
            self.status_var.set(f"Completed with errors. {len(successful_files)} succeeded, {len(failed_files)} failed.")
            error_msg = "The following files could not be processed:\n" + "\n".join(failed_files)
            messagebox.showwarning("Conversion Incomplete", error_msg)
        else:
            self.status_var.set(f"Conversion complete. All {len(successful_files)} files processed successfully.")
            
            # Show success message
            messagebox.showinfo(
                "Conversion Complete", 
                f"All files converted successfully!\n\n"
                f"{len(successful_files)} files processed and saved to:\n{self.output_dir.get()}"
            )
    
    def convert_unedited_file(self, input_file, output_file):
        """
        Convert unedited respiratory data file to edited format
        
        Parameters:
        input_file - Path to the input Excel file
        output_file - Path to save the converted file
        
        Returns:
        DataFrame with the converted data
        """
        wb = load_workbook(input_file)
        ws = wb.active

        flow_values = []
        volume_values = []
        start_collecting_flow = False
        start_collecting_volume = False
        skip_next_row_flow = False
        skip_next_row_volume = False

        # First pass to collect data
        for row in ws.iter_rows(values_only=True):
            first_column = str(row[0]).strip() if row[0] is not None else ""

            # Start collecting flow data when marker found
            if not start_collecting_flow and "ltr/s" in first_column.lower():
                skip_next_row_flow = True
                start_collecting_flow = True
                continue

            # Start collecting volume data when marker found
            if not start_collecting_volume and "ltr" == first_column.lower():
                skip_next_row_volume = True
                start_collecting_volume = True
                continue

            # Skip header rows
            if skip_next_row_flow:
                skip_next_row_flow = False
                continue

            if skip_next_row_volume:
                skip_next_row_volume = False
                continue

            # Collect flow values
            if start_collecting_flow:
                if first_column == "":
                    start_collecting_flow = False
                else:
                    try:
                        flow_values.append(float(first_column))
                    except (ValueError, TypeError):
                        pass

            # Collect volume values
            if start_collecting_volume:
                if first_column == "":
                    start_collecting_volume = False
                else:
                    try:
                        volume_values.append(float(first_column))
                    except (ValueError, TypeError):
                        pass

        # Find the maximum length between flow and volume arrays
        max_length = max(len(flow_values), len(volume_values))
        
        # Create time values (0.01s intervals) to match the longest array
        time_data = [round(0.01 * i, 2) for i in range(1, max_length + 1)]
        
        # Ensure flow and volume arrays are the same length as time by padding with NaN
        if len(flow_values) < max_length:
            flow_values.extend([np.nan] * (max_length - len(flow_values)))
        
        if len(volume_values) < max_length:
            volume_values.extend([np.nan] * (max_length - len(volume_values)))

        # Create DataFrame with all data
        processed_data = pd.DataFrame({
            "Time": time_data,
            "Vol": volume_values,
            "Flow": flow_values
        })

        # Save to new file
        processed_data.to_excel(output_file, index=False)
        return processed_data